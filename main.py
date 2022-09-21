import argparse
import os
import re
import glob
import numpy as np
import scipy.io as sio
from vis_tools import Visualizer

import torch
import torch.nn as nn
import torch.optim as optim
import model

from datasets import trainset_loader
from datasets import testset_loader
from datasets import valiset_loader
from torch.utils.data import DataLoader
from torch.autograd import Variable
import time
import openpyxl as xl
parser = argparse.ArgumentParser()
parser.add_argument("--epochs", type=int, default=200, help="number of epochs of training")
parser.add_argument("--batch_size", type=int, default=1, help="size of the batches")
parser.add_argument("--lr", type=float, default=1e-4, help="adam: learning rate")
parser.add_argument("--n_block", type=int, default=2)
parser.add_argument("--n_cpu", type=int, default=2)
parser.add_argument("--model_save_path", type=str, default="saved_models/1st")
parser.add_argument('--checkpoint_interval', type=int, default=1)

opt = parser.parse_args()
cuda = True if torch.cuda.is_available() else False
train_vis = Visualizer(env='training_regformer')

class net():
    def __init__(self):
        self.model = model.Learn(opt.n_block, views=64, dets=368, width=256, height=256, 
            dImg=0.006641*2, dDet=0.012858*2, dAng=0.006134*16, s2r=5.95, d2r=4.906, binshift=0)
        self.loss = nn.MSELoss()
        self.path = opt.model_save_path
        self.train_data = DataLoader(trainset_loader("E:\学习资料\原始数据\mayo_data_sparse_view", '64'),
            batch_size=opt.batch_size, shuffle=True, num_workers=opt.n_cpu)
        self.vali_data = DataLoader(valiset_loader("E:\学习资料\原始数据\mayo_data_sparse_view", '64'),
            batch_size=opt.batch_size*4, shuffle=False, num_workers=opt.n_cpu)   
        self.test_data = DataLoader(testset_loader("E:\学习资料\原始数据\mayo_data_sparse_view", '64'),
            batch_size=opt.batch_size*4, shuffle=False, num_workers=opt.n_cpu)             
        self.start = 0
        self.epoch = opt.epochs
        self.check_saved_model()       
        if cuda:
            self.model = self.model.cuda()
        self.optimizer = optim.AdamW([{'params':self.model.parameters(), 'initial_lr':opt.lr}], lr=opt.lr)
        self.scheduler = optim.lr_scheduler.CosineAnnealingLR(self.optimizer, T_max=200, eta_min=0, last_epoch=self.start-1)
        self.loss_file = self.path + '/loss.xlsx'
        if not os.path.exists(self.loss_file):
            wb = xl.Workbook()
            wb.save(self.loss_file)

    def check_saved_model(self):
        if not os.path.exists(self.path):
            os.makedirs(self.path)
            # self.initialize_weights()
        else:
            model_list = glob.glob(self.path + '/model_epoch_*.pth')
            if len(model_list) == 0:
                None
                # self.initialize_weights()
            else:
                last_epoch = 0
                for model in model_list:
                    epoch_num = int(re.findall(r'model_epoch_(-?[0-9]\d*).pth', model)[0])
                    if epoch_num > last_epoch:
                        last_epoch = epoch_num
                self.start = last_epoch
                self.model.load_state_dict(torch.load(
                    '%s/model_epoch_%04d.pth' % (self.path, last_epoch), map_location='cuda:0'))

    def displaywin(self, img, low=0.42, high=0.62):
        img[img<low] = low
        img[img>high] = high
        img = (img - low)/(high - low) * 255
        return img

    def initialize_weights(self):
        for module in self.model.modules():
            if isinstance(module, model.prj_module):
                module.weight.data.zero_()
            if isinstance(module, nn.Conv2d):
                nn.init.normal_(module.weight, mean=0, std=0.01)
                if module.bias is not None:
                    module.bias.data.zero_()
            if isinstance(module, nn.BatchNorm2d):
                module.weight.data.fill_(1)
                module.bias.data.zero_()

    def train(self):
        wb = xl.load_workbook(self.loss_file)
        ws = wb.active
        for epoch in range(self.start, self.epoch):
            self.model.train()
            for batch_index, data in enumerate(self.train_data):
                input_data, label_data, prj_data = data                
                if cuda:
                    input_data = input_data.cuda()
                    label_data = label_data.cuda()
                    prj_data = prj_data.cuda()
                self.optimizer.zero_grad()
                output = self.model(input_data, prj_data)
                loss = self.loss(output, label_data)
                loss.backward()
                self.optimizer.step()
                print(
                    "[Epoch %d/%d] [Batch %d/%d]: [loss: %f]"
                    % (epoch+1, self.epoch, batch_index+1, len(self.train_data), loss.item())
                )
                ws.cell(row=batch_index+epoch*len(self.train_data)+1, column=1, value=loss.item())
                train_vis.plot('Loss', loss.item())
                train_vis.img('Ground Truth', self.displaywin(label_data.detach()).cpu())
                train_vis.img('Result', self.displaywin(output.detach()).cpu())
                train_vis.img('Input', self.displaywin(input_data.detach()).cpu())
            self.model.eval()
            vali_loss = self.validate()            
            ws.cell(row=epoch+1, column=2, value=vali_loss)
            if opt.checkpoint_interval != -1 and (epoch+1) % opt.checkpoint_interval == 0:
                torch.save(self.model.state_dict(), '%s/model_epoch_%04d.pth' % (self.path, epoch+1))
            wb.save(self.loss_file)
            self.scheduler.step()

    def test(self):
        self.model.eval()
        for batch_index, data in enumerate(self.test_data):
            input_data, label_data, prj_data, res_name = data
            if cuda:
                input_data = input_data.cuda()
                label_data = label_data.cuda()
                prj_data = prj_data.cuda()
            with torch.no_grad():
                output = self.model(input_data, prj_data)
            res = output.cpu().numpy()
            output = (self.displaywin(output, low=0.0, high=1.0) / 255).view(-1,input_data.size(2),input_data.size(3)).cpu().numpy()
            label = (self.displaywin(label_data, low=0.0, high=1.0) / 255).view(-1,input_data.size(2),input_data.size(3)).cpu().numpy()
            for i in range(output.shape[0]):
                sio.savemat(res_name[i], {'data':res[i,0]})

    def validate(self):
        loss = 0
        for batch_index, data in enumerate(self.vali_data):
            input_data, label_data, prj_data = data
            if cuda:
                input_data = input_data.cuda()
                label_data = label_data.cuda()
                prj_data = prj_data.cuda()
            with torch.no_grad():
                output = self.model(input_data, prj_data)
                loss0 = self.loss(output, label_data)            
            train_vis.img('Vali_Ground Truth', self.displaywin(label_data.detach()).cpu())
            train_vis.img('Vali_Result', self.displaywin(output.detach()).cpu())
            train_vis.img('Vali_Input', self.displaywin(input_data.detach()).cpu())            
            loss += loss0.item()
        loss = loss / len(self.vali_data)
        train_vis.plot('Vali_Loss', loss)
        return loss

    def save_loss(self, loss):
        value = str(loss)
        value += "\n"
        with open(self.path + "/loss.csv", "a+") as f:
            f.write(value)

if __name__ == "__main__":
    network = net()
    network.train()
    network.test()